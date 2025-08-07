from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io
import os
from datetime import datetime, timedelta
from models import db, ShipmentEntry
from config import Config
from sqlalchemy import func, and_
from data_processor import DataProcessor

app = Flask(__name__)
app.config.from_object(Config)

# Initialize the database
db.init_app(app)
CORS(app)

# Create database tables
with app.app_context():
    db.create_all()

@app.route('/historical-data', methods=['POST'])
def get_historical_data():
    try:
        data = request.json
        start_date = data.get('startDate')
        end_date = data.get('endDate')

        # Query the database
        query = ShipmentEntry.query

        if start_date and end_date:
            query = query.filter(
                and_(
                    func.date(ShipmentEntry.flight_date) >= start_date,
                    func.date(ShipmentEntry.flight_date) <= end_date
                )
            )

        # Execute query and convert to list of dictionaries
        entries = query.all()
        results = [entry.to_dict() for entry in entries]

        return jsonify({
            'data': results,
            'total_records': len(results),
            'results': {
                'china_post': {
                    'available': True,
                    'records_processed': len(results)
                },
                'cbp': {
                    'available': True,
                    'records_processed': len(results)
                }
            }
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Template file paths
CP_TEMPLATE = "templates/China Post data source file template.xlsx"
CBP_TEMPLATE = "templates/CBP transported package worksheet file template.xlsx"

# IODA data file path
IODA_DATA_FILE = "../../data processing/Sample_Data_from_IODA_v2 (China Post).xlsx"

# Column mappings from original code
CP_COLUMNS = [
    "*运单号 (AWB Number)",
    "*始发站（Departure station）",
    "*目的站（Destination）",
    "*件数(Pieces)",
    "*重量 (Weight)",
    "航司(Airline)",
    "航班号 (Flight Number)",
    "航班日期 (Flight Date)",
    "一个航班的邮件item总数 (Total mail items per flight)",
    "一个航班的邮件总重量 (Total mail weight per flight)",
    "*运价类型 (Rate Type)",
    "*费率 (Rate)",
    "*航空运费 (Air Freight)",
    "代理人的其他费用 (Agent's Other Charges)",
    "承运人的其他费用 (Carrier's Other Charges)",
    "*总运费 (Total Charges)"
]

CBP_COLUMNS = [
    "Carrier Code",
    "Flight/ Trip Number",
    "Tracking Number",
    "Arrival Port Code",
    "Arrival Date",
    "Declared Value (USD)"
]

# CBP header mapping
CBP_HEADER_LOOKUP = {
    "Carrier Code": "Carrier Code",
    "Flight/ Trip Number": "Flight/ Trip Number", 
    "Tracking Number": "Tracking Number",
    "Arrival Port Code": "Arrival Port Code",
    "Arrival Date": "Arrival Date",
    "Declared Value (USD)": "Declared Value (USD)"
}

def build_header_map(ws, header_rows):
    """Helper function to map template headers to column indices"""
    mapping = {}
    for row in header_rows:
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                mapping[cell.value.strip()] = cell.column
    return mapping

def filter_cp_data(df):
    """Filter out instruction/sample rows from China Post data"""
    instruction_keywords = [
        "此行开始填具体某个邮件的信息",
        "Below shows sample data for individual mail items",
        "航司提供",
        "To be completed by airline",
        "根据启运口岸提供固定地址",
        "Fixed sender address in English",
        "根据目的口岸提供固定地址", 
        "Fixed recipient address in English",
        "CP将要求所有美国路向的邮件收寄时必须按照美元申报",
        "CP will require all US-bound mail to declare value in USD"
    ]
    
    mask = df["*运单号 (AWB Number)"].astype(str).apply(
        lambda x: not any(keyword in str(x) for keyword in instruction_keywords)
    )
    return df[mask].reset_index(drop=True)

def save_workflow_output_to_database(internal_df):
    """Save workflow output (Internal Use format) to database"""
    new_entries = 0
    skipped_entries = 0
    
    for _, row in internal_df.iterrows():
        # Check if entry already exists
        tracking_number = str(row.get('Tracking Number', ''))
        receptacle_id = str(row.get('Receptacle', ''))
        flight_number = str(row.get('Flight Number 1', ''))
        flight_date = str(row.get('Flight Date 1', ''))
        
        existing_entry = ShipmentEntry.query.filter_by(
            tracking_number=tracking_number,
            receptacle_id=receptacle_id,
            flight_number=flight_number,
            flight_date=flight_date
        ).first()
        
        if existing_entry:
            skipped_entries += 1
            continue  # Skip duplicate entry
            
        # Create new shipment entry with complete workflow data
        entry = ShipmentEntry(
            # Core identification
            pawb=str(row.get('PAWB', '')),
            cardit=str(row.get('CARDIT', '')),
            tracking_number=tracking_number,
            receptacle_id=receptacle_id,
            
            # Flight information
            host_origin_station=str(row.get('Host Origin Station', '')),
            host_destination_station=str(row.get('Host Destination Station', '')),
            flight_carrier_1=str(row.get('Flight Carrier 1', '')),
            flight_number_1=str(row.get('Flight Number 1', '')),
            flight_date_1=str(row.get('Flight Date 1', '')),
            flight_carrier_2=str(row.get('Flight Carrier 2', '')),
            flight_number_2=str(row.get('Flight Number 2', '')),
            flight_date_2=str(row.get('Flight Date 2', '')),
            arrival_date=str(row.get('Arrival Date', '')),
            uld_number=str(row.get('ULD Number', '')),
            
            # Package details
            bag_weight=str(row.get('Bag Weight', '')),
            bag_number=str(row.get('Bag Number', '')),
            packets_in_receptacle=str(row.get('Number of packets inside Receptacle', '')),
            
            # Content and customs
            declared_content=str(row.get('Declared content', '')),
            hs_code=str(row.get('HS Code', '')),
            declared_value=str(row.get('Declared Value', '')),
            currency=str(row.get('Currency', '')),
            tariff_amount=str(row.get('Tariff amount', '')),
            
            # CBP fields
            carrier_code=str(row.get('Flight Carrier 1', '')),  # Use carrier for CBP
            arrival_port_code='4701',  # Default port code
            declared_value_usd=str(row.get('Declared Value', '')),  # Assuming USD
            
            # Legacy fields for backward compatibility
            awb_number=str(row.get('PAWB', '')),
            departure_station=str(row.get('Host Origin Station', '')),
            destination=str(row.get('Host Destination Station', '')),
            weight=str(row.get('Bag Weight', '')),
            airline=str(row.get('Flight Carrier 1', '')),
            flight_number=str(row.get('Flight Number 1', '')),
            flight_date=str(row.get('Flight Date 1', '')),
            total_charges=str(row.get('Tariff amount', ''))
        )
        
        db.session.add(entry)
        new_entries += 1
    
    return new_entries, skipped_entries

def create_china_post_output(df):
    """Create China Post output Excel file and save to database"""
    # Filter data
    cp_df = df[CP_COLUMNS].copy()
    cp_df = filter_cp_data(cp_df)
    
    # Load template to extract headers
    wb_cp_template = load_workbook(CP_TEMPLATE)
    ws_cp_template = wb_cp_template.active
    
    # Create new workbook
    wb_cp = Workbook()
    ws_cp = wb_cp.active
    
    # Copy header rows (1-2) from template
    for row_num in [1, 2]:
        for col_num in range(1, ws_cp_template.max_column + 1):
            template_cell = ws_cp_template.cell(row=row_num, column=col_num)
            new_cell = ws_cp.cell(row=row_num, column=col_num)
            new_cell.value = template_cell.value
            # Copy header formatting
            if template_cell.font:
                new_cell.font = Font(name=template_cell.font.name, size=template_cell.font.size, bold=template_cell.font.bold)
            if template_cell.alignment:
                new_cell.alignment = Alignment(horizontal=template_cell.alignment.horizontal, vertical=template_cell.alignment.vertical, wrap_text=template_cell.alignment.wrap_text)
    
    # Set header row heights
    ws_cp.row_dimensions[1].height = 51.0
    ws_cp.row_dimensions[2].height = 81.0
    
    # Build header mapping
    header_map_cp = build_header_map(ws_cp, ws_cp.iter_rows(min_row=1, max_row=2))
    
    # Column header lookup
    cp_header_lookup = {
        "*运单号 (AWB Number)": "*运单号 \n(AWB Number)",
        "*始发站（Departure station）": "*始发站\n（Departure station）",
        "*目的站（Destination）": "*目的站\n（Destination）",
        "*件数(Pieces)": "*件数(Pieces)",
        "*重量 (Weight)": "*重量 \n(Weight) ",
        "航司(Airline)": "航司(Airline)",
        "航班号 (Flight Number)": "航班号\n(Flight Number)",
        "航班日期 (Flight Date)": "航班日期\n(Flight Date)",
        "一个航班的邮件item总数 (Total mail items per flight)": "一个航班的邮件item总数\n (Total mail items per flight)",
        "一个航班的邮件总重量 (Total mail weight per flight)": "一个航班的邮件总重量\n(Total mail weight per flight)",
        "*运价类型 (Rate Type)": "*运价类型\n (Rate Type)",
        "*费率 (Rate)": "*费率\n (Rate)",
        "*航空运费 (Air Freight)": "*航空运费\n(Air Freight)",
        "代理人的其他费用 (Agent's Other Charges)": "代理人的其他费用\n(Agent's Other Charges)",
        "承运人的其他费用 (Carrier's Other Charges)": "承运人的其他费用\n (Carrier's Other Charges)",
        "*总运费 (Total Charges)": "*总运费\n (Total Charges)"
    }
    
    # Insert data starting from row 3
    start_row_cp = 3
    standard_height = 15
    default_font = Font(name="Calibri", size=11)
    default_alignment = Alignment(horizontal="left", vertical="center")
    
    for i, row in cp_df.iterrows():
        excel_row = start_row_cp + i
        for df_col, template_header in cp_header_lookup.items():
            col_idx = header_map_cp.get(template_header)
            if col_idx:
                cell = ws_cp.cell(row=excel_row, column=col_idx, value=row[df_col])
                cell.font = default_font
                cell.alignment = default_alignment
        
        ws_cp.row_dimensions[excel_row].height = standard_height
    
    # Save to BytesIO
    output = io.BytesIO()
    wb_cp.save(output)
    output.seek(0)
    
    return output

def create_cbp_output(df):
    """Create CBP output Excel file and save to database"""
    # Extract CBP data
    cbp_df = df[CBP_COLUMNS].copy()
    
    # Load CBP template
    wb_cbp_template = load_workbook(CBP_TEMPLATE)
    ws_cbp_template = wb_cbp_template.active
    
    # Create new CBP workbook
    wb_cbp = Workbook()
    ws_cbp = wb_cbp.active
    
    # Copy header rows (1-3) from template
    for row_num in [1, 2, 3]:
        for col_num in range(1, ws_cbp_template.max_column + 1):
            template_cell = ws_cbp_template.cell(row=row_num, column=col_num)
            new_cell = ws_cbp.cell(row=row_num, column=col_num)
            new_cell.value = template_cell.value
            # Copy header formatting
            if template_cell.font:
                new_cell.font = Font(name=template_cell.font.name, size=template_cell.font.size, bold=template_cell.font.bold)
            if template_cell.alignment:
                new_cell.alignment = Alignment(horizontal=template_cell.alignment.horizontal, vertical=template_cell.alignment.vertical, wrap_text=template_cell.alignment.wrap_text)
    
    # Set header row heights
    ws_cbp.row_dimensions[1].height = 15.4
    ws_cbp.row_dimensions[2].height = 15.4
    ws_cbp.row_dimensions[3].height = 15.75
    
    # Build header mapping
    header_map_cbp = build_header_map(ws_cbp, [ws_cbp[3]])
    
    # Insert data starting from row 4
    start_row_cbp = 4
    standard_height = 15
    default_font = Font(name="Calibri", size=11)
    default_alignment = Alignment(horizontal="left", vertical="center")
    
    for i, row in cbp_df.iterrows():
        excel_row = start_row_cbp + i
        for col_name in CBP_COLUMNS:
            template_header = CBP_HEADER_LOOKUP.get(col_name, col_name)
            col_idx = header_map_cbp.get(template_header)
            if col_idx:
                cell = ws_cbp.cell(row=excel_row, column=col_idx, value=row[col_name])
                cell.font = default_font
                cell.alignment = default_alignment
        
        ws_cbp.row_dimensions[excel_row].height = standard_height
    
    # Save to BytesIO
    output = io.BytesIO()
    wb_cbp.save(output)
    output.seek(0)
    
    return output

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    # Also include database record count for debugging
    record_count = ShipmentEntry.query.count()
    return jsonify({
        "status": "healthy", 
        "timestamp": datetime.now().isoformat(),
        "database_records": record_count
    })

@app.route('/upload-cnp-file', methods=['POST'])
def upload_cnp_file():
    """Upload and process raw CNP Excel file"""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        # Check file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({"error": "Invalid file format. Please upload an Excel file."}), 400
        
        # Save uploaded file temporarily
        temp_path = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        file.save(temp_path)
        
        try:
            # Read the raw CNP data from the first sheet (header=None for custom parsing)
            cnp_df = pd.read_excel(temp_path, sheet_name='Raw data provided by CNP', header=None)
            
            # Initialize data processor
            processor = DataProcessor(IODA_DATA_FILE)
            
            # Process the data to get both internal use and CBP formats
            internal_df, cbp_df = processor.process_cnp_data(cnp_df)
            
            # Get the merged IODA data for database storage
            processed_ioda_data = processor.get_processed_data()
            
            # Save workflow output to database
            if internal_df is not None and not internal_df.empty:
                new_entries, skipped_entries = save_workflow_output_to_database(internal_df)
                db.session.commit()
                print(f"Saved to database: {new_entries} new entries, {skipped_entries} duplicates skipped")
            else:
                new_entries, skipped_entries = 0, 0
            
            return jsonify({
                "success": True,
                "message": "CNP file processed successfully",
                "results": {
                    "internal_use": {
                        "available": True,
                        "records_processed": len(internal_df) if not internal_df.empty else 0
                    },
                    "cbp": {
                        "available": True,
                        "records_processed": len(cbp_df) if not cbp_df.empty else 0
                    }
                },
                "total_records": len(internal_df) if internal_df is not None else 0,
                "new_entries": new_entries,
                "skipped_duplicates": skipped_entries
            })
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    except Exception as e:
        # Clean up temporary file in case of error
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"error": str(e)}), 500

@app.route('/process-data', methods=['POST'])
def process_data():
    """Process input data and generate both output files"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Validate required columns exist
        missing_cp_cols = [col for col in CP_COLUMNS if col not in df.columns]
        missing_cbp_cols = [col for col in CBP_COLUMNS if col not in df.columns]
        
        if missing_cp_cols and missing_cbp_cols:
            return jsonify({
                "error": "Missing required columns",
                "missing_cp_columns": missing_cp_cols,
                "missing_cbp_columns": missing_cbp_cols
            }), 400
        
        # Generate both files
        results = {}
            
        if not missing_cp_cols:
            cp_output = create_china_post_output(df)
            results["china_post"] = {
                "available": True,
                "records_processed": len(df[CP_COLUMNS])
            }
        else:
            results["china_post"] = {
                "available": False,
                "missing_columns": missing_cp_cols
            }
        
        if not missing_cbp_cols:
            cbp_output = create_cbp_output(df)
            results["cbp"] = {
                "available": True,
                "records_processed": len(df[CBP_COLUMNS])
            }
        else:
            results["cbp"] = {
                "available": False,
                "missing_columns": missing_cbp_cols
            }
        
        # This endpoint is for legacy JSON data processing
        # For CNP file uploads, data is saved in the upload-cnp-file endpoint
        new_entries = 0
        skipped_entries = 0

        return jsonify({
            "success": True,
            "message": "Data processed successfully",
            "results": results,
            "total_records": len(df),
            "new_entries": new_entries,
            "skipped_duplicates": skipped_entries
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/generate-china-post', methods=['POST'])
def generate_china_post():
    """Generate China Post (Internal Use) output file"""
    try:
        # Check if we're getting processed data or need to process a file
        if 'file' in request.files:
            # Process uploaded file
            file = request.files['file']
            if file.filename == '':
                return jsonify({"error": "No file selected"}), 400
            
            # Save and process file
            temp_path = f"temp_cp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
            file.save(temp_path)
            
            try:
                cnp_df = pd.read_excel(temp_path, sheet_name='Raw data provided by CNP', header=None)
                processor = DataProcessor(IODA_DATA_FILE)
                internal_df, cbp_df = processor.process_cnp_data(cnp_df)
                
                if internal_df is None or internal_df.empty:
                    return jsonify({"error": "No Internal Use data processed"}), 400
                
                # Create Excel output with Internal Use format
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    internal_df.to_excel(writer, sheet_name='Internal Use Output', index=False)
                output.seek(0)
                
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        else:
            # Use existing JSON data processing for backward compatibility
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400
            
            df = pd.DataFrame(data)
            
            # Check for required columns (using old format for compatibility)
            missing_cols = [col for col in CP_COLUMNS if col not in df.columns]
            if missing_cols:
                return jsonify({
                    "error": "Missing required columns for China Post output",
                    "missing_columns": missing_cols
                }), 400
            
            # Generate file using old method
            output = create_china_post_output(df)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"internal_use_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/generate-cbp', methods=['POST'])
def generate_cbp():
    """Generate CBP output file"""
    try:
        # Check if we're getting processed data or need to process a file
        if 'file' in request.files:
            # Process uploaded file
            file = request.files['file']
            if file.filename == '':
                return jsonify({"error": "No file selected"}), 400
            
            # Save and process file
            temp_path = f"temp_cbp_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
            file.save(temp_path)
            
            try:
                cnp_df = pd.read_excel(temp_path, sheet_name='Raw data provided by CNP', header=None)
                processor = DataProcessor(IODA_DATA_FILE)
                internal_df, cbp_df = processor.process_cnp_data(cnp_df)
                
                if cbp_df is None or cbp_df.empty:
                    return jsonify({"error": "No CBP data processed"}), 400
                
                # Create Excel output with CBP format
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    cbp_df.to_excel(writer, sheet_name='CBP Output', index=False)
                output.seek(0)
                
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        else:
            # Use existing JSON data processing for backward compatibility
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided"}), 400
            
            df = pd.DataFrame(data)
            
            # Check for required columns (using old format for compatibility)
            missing_cols = [col for col in CBP_COLUMNS if col not in df.columns]
            if missing_cols:
                return jsonify({
                    "error": "Missing required columns for CBP output",
                    "missing_columns": missing_cols
                }), 400
            
            # Generate file using old method
            output = create_cbp_output(df)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"cbp_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/columns', methods=['GET'])
def get_columns():
    """Get the required column definitions"""
    return jsonify({
        "china_post_columns": CP_COLUMNS,
        "cbp_columns": CBP_COLUMNS,
        "total_unique_columns": list(set(CP_COLUMNS + CBP_COLUMNS))
    })

@app.route('/delete-records', methods=['DELETE'])
def delete_records():
    """Delete multiple shipment records by IDs"""
    try:
        data = request.json
        record_ids = data.get('ids', [])
        
        if not record_ids:
            return jsonify({"error": "No record IDs provided"}), 400
        
        # Delete records
        deleted_count = 0
        for record_id in record_ids:
            entry = ShipmentEntry.query.get(record_id)
            if entry:
                db.session.delete(entry)
                deleted_count += 1
        
        db.session.commit()
        
        return jsonify({
            "message": f"Successfully deleted {deleted_count} records",
            "deleted_count": deleted_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/update-record', methods=['PUT'])
def update_record():
    """Update a single shipment record"""
    try:
        data = request.json
        record_id = data.get('id')
        updates = data.get('updates', {})
        
        if not record_id:
            return jsonify({"error": "No record ID provided"}), 400
        
        if not updates:
            return jsonify({"error": "No updates provided"}), 400
        
        # Find the record
        entry = ShipmentEntry.query.get(record_id)
        if not entry:
            return jsonify({"error": "Record not found"}), 404
        
        # Map frontend field names to database field names
        field_mapping = {
            '*运单号 (AWB Number)': 'awb_number',
            '*始发站（Departure station）': 'departure_station',
            '*目的站（Destination）': 'destination',
            '*件数(Pieces)': 'pieces',
            '*重量 (Weight)': 'weight',
            '航司(Airline)': 'airline',
            '航班号 (Flight Number)': 'flight_number',
            '航班日期 (Flight Date)': 'flight_date',
            '一个航班的邮件item总数 (Total mail items per flight)': 'total_mail_items',
            '一个航班的邮件总重量 (Total mail weight per flight)': 'total_mail_weight',
            '*运价类型 (Rate Type)': 'rate_type',
            '*费率 (Rate)': 'rate',
            '*航空运费 (Air Freight)': 'air_freight',
            "代理人的其他费用 (Agent's Other Charges)": 'agent_charges',
            "承运人的其他费用 (Carrier's Other Charges)": 'carrier_charges',
            '*总运费 (Total Charges)': 'total_charges',
            'Carrier Code': 'carrier_code',
            'Flight/ Trip Number': 'flight_number',
            'Tracking Number': 'tracking_number',
            'Arrival Port Code': 'arrival_port_code',
            'Arrival Date': 'arrival_date',
            'Declared Value (USD)': 'declared_value_usd'
        }
        
        # Update the record
        updated_fields = []
        for frontend_field, value in updates.items():
            db_field = field_mapping.get(frontend_field)
            if db_field and hasattr(entry, db_field):
                setattr(entry, db_field, str(value))
                updated_fields.append(frontend_field)
        
        db.session.commit()
        
        return jsonify({
            "message": f"Successfully updated record",
            "updated_fields": updated_fields,
            "record": entry.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/routes', methods=['GET'])
def get_routes():
    """Get all unique route combinations from the database"""
    try:
        # Query distinct departure_station and destination combinations
        routes_query = db.session.query(
            ShipmentEntry.departure_station,
            ShipmentEntry.destination
        ).filter(
            and_(
                ShipmentEntry.departure_station.isnot(None),
                ShipmentEntry.destination.isnot(None),
                ShipmentEntry.departure_station != '',
                ShipmentEntry.destination != ''
            )
        ).distinct().all()
        
        # Convert to list of route objects
        routes = []
        for departure, destination in routes_query:
            routes.append({
                'origin': departure,
                'destination': destination,
                'route_id': f"{departure}->{destination}"
            })
        
        return jsonify({
            'routes': routes,
            'total_routes': len(routes)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)